import telebot
from telebot import types
from openpyxl import load_workbook
import text
import datetime

TOKEN = '6214187356:AAFpJfITT4_D3mXvpEW1mb1XMWGEdawU_YI'
bot = telebot.TeleBot(TOKEN)

#Создание словоря с iD и выбранным языком
user_languages = {}
#Команда /start - выбор языка
@bot.message_handler(commands=['start'])
def language_selection(message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text=text.language_selection)
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "🇺🇦Українська")
    btn2 = types.KeyboardButton(text ="🇬🇧English")
    kb.add(btn1, btn2,)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
    bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
# Функция для обработки выбора языка пользователем
@bot.message_handler(func=lambda message: message.text in ['🇬🇧English', '🇺🇦Українська'])
def language_preservation(message):
    #Cохраняем выбранный язык в словаре
    user_languages[message.chat.id] = message.text
    #Приветствие
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/start.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    main_menu (message)
# /share (Поделиться)
@bot.message_handler(commands=['share'])#Не переводиться на англ.
def share (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_zag_share, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_osn_share, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.zag_share, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.osn_share, parse_mode='HTML')
# Меню - Главное + команда
@bot.message_handler(commands=['menu']) #Не переводиться на англ.
def main_menu (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_privetstvie)
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f64fHelp the project")
        btn2 = types.KeyboardButton(text="\U0001faf6About us")
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        bot.send_message(message.chat.id, text = text.privetstvie)
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f198Отримати допомогу")
        btn2 = types.KeyboardButton(text ="\U0001f64fДопомогти проекту")
        btn3 = types.KeyboardButton(text ="\U0001f3ebОсвітні заходи")
        btn4 = types.KeyboardButton(text="\U0001faf6Про нас")
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#------------ Меню Отримати допомогу-----
#Меню Отримати допомогу
@bot.message_handler(func=lambda message: message.text == "\U0001f198Отримати допомогу")
def main_help_project (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "👨‍⚖️Юридична консультація")
    btn2 = types.KeyboardButton(text = "📦Гуманітарна допомога")
    btn3 = types.KeyboardButton(text = "🙏Реалізувати Вашу мрію")
    btn4 = types.KeyboardButton(text = "\U0001fa96Допомога для ЗСУ")
    btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
    kb.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Кнопки возврата в главное меню   
def button_back_main_help_project (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "\U0001f519Hазад")
    btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")
    kb.add(btn1, btn2)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
# Меню - Юридична консультація    
@bot.message_handler(func=lambda message: message.text == "👨‍⚖️Юридична консультація")
def legal_consultation(message):
    global user_choice
    user_choice = {}
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.legal_consultation, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1 = types.InlineKeyboardButton(text='\U0001fa96Я військовослужбовець', callback_data='Військовослужбовець')
    btn2 = types.InlineKeyboardButton(text='🧳Я внутрішньо переміщена особа', callback_data='Внутрішньо переміщена особа')
    btn3 = types.InlineKeyboardButton(text='😢Я людина, яка постраждала від війни', callback_data='Людина, яка постраждала від війни')
    btn4 = types.InlineKeyboardButton(text='🤲Я волонтер', callback_data='Волонтер')
    kb.add(btn1, btn2, btn3, btn4)
    bot.send_message(message.chat.id, "💁🏻‍♂️Тут ви можете отримати допомогу, але вам потрібно розповісти про себе", reply_markup=kb)
    button_back_main_help_project (message)
#Фукция проверки повторного нажатия инлайн кнопки
@bot.callback_query_handler(func=lambda call: True)
def callback_query(call):
    global var_button_legal
    global var_button
    chat_id = call.message.chat.id
    # Проверяем, был ли выбран ответ для данного пользователя
    if chat_id in user_choice:
        bot.send_message(chat_id, text=text.in_button, parse_mode='HTML')
        return
    # Сохраняем выбранный ответ для данного пользователя
    user_choice[chat_id] = call.data
    # Ваш код обработки выбора
    bot.answer_callback_query(callback_query_id=call.id)
    var_button = call.data
    if call.data != None:
        var_button_legal = '👨‍⚖️Юридична консультація'
        sent = bot.send_message(chat_id, text=text.your_situation, parse_mode='HTML')
        bot.register_next_step_handler(sent, ignor_button_help_project)
    else:
        legal_consultation(call.message)
#Отправка в Эксельку
def humanitarian_dream_help_zsy(message,var_button_legal,var_button=None):
    # Загружаем эксельку
    wb = load_workbook('request.xlsx')
    # Открываем
    sheet = wb.active
    # Находим последнюю строку с данными
    last_row = sheet.max_row + 1
    #Текущее время
    yest_datetime = datetime.datetime.now()
    # Добавляем новые данные в последнюю строку
    sheet.cell(row=last_row, column=1, value=yest_datetime)
    sheet.cell(row=last_row, column=2, value=var_button_legal)
    sheet.cell(row=last_row, column=3, value=var_button)
    sheet.cell(row=last_row, column=4, value=message.text)
    sheet.cell(row=last_row, column=5, value=message.from_user.first_name)
    sheet.cell(row=last_row, column=6, value=message.from_user.last_name)
    sheet.cell(row=last_row, column=7, value=message.from_user.username)
    sheet.cell(row=last_row, column=8, value=message.chat.id)
    # Сохраняем изменения в файл
    wb.save('request.xlsx') 
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text=text.thank_contacting, parse_mode='HTML')
    bot.send_message(message.chat.id, text=text.button_driver)
    #Отправка в ТГ канал уведомления
    bot.send_message('-1001801043894', "Вам повідомлення: \U0001f198Отримати допомогу (request)")
#Обработка, чтобы не отправлялись кнопки в сообщениях
def ignor_button_help_project(message):
    if message.text == '/start':
        language_selection(message)
    elif message.text == '/menu':
        main_menu(message)
    elif message.text == '/share':
        share(message)
        #Назад через символ H (англ)
    elif message.text == '\U0001f519Hазад':
        main_help_project (message)
    elif message.text == '\u23EAВ головне меню':
        main_menu(message)
        # Проигнорировать сообщения, которые не являются текстом!
    elif message.content_type != 'text':
        bot.send_chat_action(message.chat.id, 'typing')
        bot.send_message(message.chat.id, text=text.get_help_not_understand, parse_mode='HTML')
    else:
        humanitarian_dream_help_zsy(message,var_button_legal,var_button)
#Гумонітарна допомога
@bot.message_handler(func=lambda message: message.text == "📦Гуманітарна допомога")
def Humanitarian_aid (message):
    global var_button
    var_button = None
    global var_button_legal
    var_button_legal = '📦Гуманітарна допомога'
    bot.send_chat_action(message.chat.id, 'typing')
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button_help_project)
    button_back_main_help_project (message)
#Реалізувати Вашу мрію
@bot.message_handler(func=lambda message: message.text == "🙏Реалізувати Вашу мрію")
def realize_your_dream(message):
    global var_button
    var_button = None
    global var_button_legal
    var_button_legal = '🙏Реалізувати Вашу мрію'
    bot.send_chat_action(message.chat.id, 'typing')
    sent = bot.send_message(message.chat.id, text = text.realize_your_dream, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button_help_project)
    button_back_main_help_project (message)
#Допомога для ЗСУ
@bot.message_handler(func=lambda message: message.text == "\U0001fa96Допомога для ЗСУ")
def help_zsy(message):
    global var_button
    var_button = None
    global var_button_legal
    var_button_legal = '\U0001fa96Допомога для ЗСУ'
    bot.send_chat_action(message.chat.id, 'typing')
    sent = bot.send_message(message.chat.id, text = text.help_zsy, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button_help_project)
    button_back_main_help_project (message)
#------------ конец----Меню Отримати допомогу-----

#------------ Меню Допомогти проекту-----
@bot.message_handler(func=lambda message: message.text == "\U0001f64fДопомогти проекту" or message.text =="\U0001f64fHelp the project")
def main_menu_donats (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "💰Financial assistance")
        btn2 = types.KeyboardButton(text = "🧦Other assistance")
        btn3 = types.KeyboardButton(text = "❓What will your help be aimed to")
        btn4 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "💰Фінансова допомога")
        btn2 = types.KeyboardButton(text = "🧦Інша допомога")
        btn3 = types.KeyboardButton(text = "❓На що буде спрямована ваша допомога")
        btn4 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата в Допомогти проекту
def menu_vozvrata_donats (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Bаck")
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Нaзад")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата в Монобанка,Банк.рах.,кріпта
def menu_vozvrata_mono_krypto (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Back")  
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Haзад")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню 💰Фінансова допомога-----
@bot.message_handler(func=lambda message: message.text == "💰Фінансова допомога" or message.text =="💰Financial assistance")
def menu_money (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001fad9Monobanka")
        btn2 = types.KeyboardButton(text = "💳Bank account")
        btn3 = types.KeyboardButton(text = "\U0001f519Bаck")
        btn4 = types.KeyboardButton(text = "\U0001fa99Cryptocurrency")
        btn5 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001fad9Monobanka")
        btn2 = types.KeyboardButton(text = "💳Банківські рахунки")
        btn3 = types.KeyboardButton(text = "\U0001f519Нaзад")
        btn4 = types.KeyboardButton(text = "\U0001fa99Криптовалюта")  
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню --Моно банка
@bot.message_handler(func=lambda message: message.text == "\U0001fad9Monobanka")
def menu_mono (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/monobanka.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='\U0001fad9Monobanka', url='https://send.monobank.ua/jar/AMrMtk2Vz')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text.eng_follow, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text.follow, reply_markup = kb)
    menu_vozvrata_mono_krypto (message)
#Меню 💳Банківські рахунки
@bot.message_handler(func=lambda message: message.text == "💳Банківські рахунки" or message.text =="💳Bank account")
def bank_accounts (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f1fa\U0001f1f8USD")
        btn2 = types.KeyboardButton(text = "\U0001f1ea\U0001f1faEUR")  
        btn3 = types.KeyboardButton(text = "\U0001f519Back")
        btn4 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f1fa\U0001f1e6UAN")
        btn2 = types.KeyboardButton(text = "\U0001f1fa\U0001f1f8USD")
        btn3 = types.KeyboardButton(text = "\U0001f519Haзад")
        btn4 = types.KeyboardButton(text = "\U0001f1ea\U0001f1faEUR")  
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата с 💳Банківські рахунки
def menu_vozvrata_bank_accounts (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Baсk")  
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Назaд")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
# Кнопка UAN
@bot.message_handler(func=lambda message: message.text == "\U0001f1fa\U0001f1e6UAN")
def button_uan (message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.uan, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка USD
@bot.message_handler(func=lambda message: message.text == "\U0001f1fa\U0001f1f8USD")
def button_uan (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_usd, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.usd, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка EUR
@bot.message_handler(func=lambda message: message.text == "\U0001f1ea\U0001f1faEUR")
def button_uan (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_eur, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.eur, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка - Криптовалюта
@bot.message_handler(func=lambda message: message.text == "\U0001fa99Криптовалюта" or message.text == "\U0001fa99Cryptocurrency")
def crypto (message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.crypto_check, parse_mode='HTML')
    menu_vozvrata_mono_krypto (message)
#Кнопка ❓На що буде спрямована ваша допомога
@bot.message_handler(func=lambda message: message.text == "❓На що буде спрямована ваша допомога" or message.text == "❓What will your help be aimed to")
def your_help_is_straightened (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_what_money_for, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.what_money_for, parse_mode='HTML')
    menu_vozvrata_donats (message)
#🧦Інша допомога
@bot.message_handler(func=lambda message: message.text == "🧦Інша допомога" or message.text == "🧦Other assistance")
def other_help (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        sent = bot.send_message(message.chat.id, text = text.eng_other_help_t, parse_mode='HTML')
    else:
        sent = bot.send_message(message.chat.id, text = text.other_help_t, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button_other_help)
    menu_vozvrata_donats (message)
#Обработка, чтобы не отправлялись кнопки в сообщениях
def ignor_button_other_help(message):
    if message.text == '/start':
        language_selection(message)
    elif message.text == '/menu':
        main_menu(message)
    elif message.text == '/share':
        share(message)
    #Назад с направления Допомгти проекту (первый символ анг. а)
    elif message.text == '\U0001f519Нaзад':
        main_menu_donats (message)
    elif message.text == '\u23EAВ головне меню':
        main_menu(message)
    elif message.text == "\U0001f519Bаck":
        main_menu_donats (message)
    elif message.text == "\u23EATo main menu":
        main_menu(message)
    elif message.content_type != 'text':
        bot.send_chat_action(message.chat.id, 'typing')
        chat_id = message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(message.chat.id, text=text.eng_help_project_not_understand, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, text=text.help_project_not_understand, parse_mode='HTML')
        other_help (message)
    else:
        other_help_excel(message)
#Сохранние в эксельку
def other_help_excel (message):
    wb = load_workbook('proposal.xlsx')
    # Открываем
    sheet = wb.active
    # Находим последнюю строку с данными
    last_row = sheet.max_row + 1
    #Текущее время
    yest_datetime = datetime.datetime.now()
    # Добавляем новые данные в последнюю строку
    sheet.cell(row=last_row, column=1, value=yest_datetime)
    sheet.cell(row=last_row, column=2, value=message.text)
    sheet.cell(row=last_row, column=3, value=message.from_user.first_name)
    sheet.cell(row=last_row, column=4, value=message.from_user.last_name)
    sheet.cell(row=last_row, column=5, value=message.from_user.username)
    sheet.cell(row=last_row, column=6, value=message.chat.id)
    # Сохраняем изменения в файл
    wb.save('proposal.xlsx') 
    # bot.send_message(message.chat.id, text=text.thank_contacting)
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_thank_contacting, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_button_driver)
    #Отправка в ТГ канал уведомления
        bot.send_message('-1001801043894', "Вам повідомлення: \U0001f64fHelp the project (proposal)")
    else:
        bot.send_message(message.chat.id, text=text.thank_contacting, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.button_driver)
    #Отправка в ТГ канал уведомления
        bot.send_message('-1001801043894', "Вам повідомлення: \U0001f64fДопомогти проекту (proposal)")
#------------ конец----Меню Допомогти проекту-----

#------------Меню Освітні заходи
@bot.message_handler(func=lambda message: message.text == "\U0001f3ebОсвітні заходи")
def educational_activities(message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.educational_activities, parse_mode='HTML')
#------------ конец----Освітні заходи-----

#------------ Меню - Про нас
@bot.message_handler(func=lambda message: message.text == "\U0001faf6Про нас" or message.text == "\U0001faf6About us")
def menu_about_us (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text="🧑‍💻On development stage")
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "🧑‍💼Наші засновники")
        btn2 = types.KeyboardButton(text = "👪Наша команда")
        btn3 = types.KeyboardButton(text = "🥇Наші досягнення")
        btn4 = types.KeyboardButton(text = "💬Ми в соціальних мережах")
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")
        kb.add(btn1, btn2, btn3, btn4,btn5)
        bot.send_message(message.chat.id, text=text.button_driver, reply_markup=kb)
#Кнопки возврата в меню: Про нас и Главное меню
def button_back_about_us (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "\U0001f519Назад")
    btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")
    kb.add(btn1, btn2)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Контент - Основатели
@bot.message_handler(func=lambda message: message.text == "🧑‍💼Наші засновники")
def osnovatel (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/team/gatesh.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.vadim_gatezh, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/vadim.gatezh')
    btn2= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/v_gatezh')
    btn3= types.InlineKeyboardButton(text='Telegram channel', url='https://t.me/v_gatezh_novyny')
    kb.add(btn1, btn2, btn3)
    bot.send_message(message.chat.id, "Соціальна мережа:", reply_markup = kb)
    
    photo = open('image/team/visotskiy.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.pavlo_vysotsky, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/profile.php?id=100015205090408')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа:", reply_markup = kb)
    button_back_about_us (message)
#Контент - Команда
@bot.message_handler(func=lambda message: message.text == "👪Наша команда")
def team (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/team/mironuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.anna_mironyuk, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/anya.myroniuk')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа", reply_markup = kb)

    photo = open('image/team/merezhuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.anastasia_merezhuk, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/merranst')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа", reply_markup = kb)

    photo = open('image/team/semenchuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.anastasia_semenchuk, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/nastia_semenchuk')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа", reply_markup = kb)

    photo = open('image/team/birkova.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.julia_birkova, parse_mode='HTML')

    photo = open('image/team/bagirov.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.vagif_bagirov, parse_mode='HTML')

    photo = open('image/team/shaporda.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.anastasia_shaporda, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/shaporda.design')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа", reply_markup = kb)

    photo = open('image/team/torska.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.khrystyna_torska, parse_mode='HTML')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/kristorska')
    kb.add(btn1)
    bot.send_message(message.chat.id, "Соціальна мережа", reply_markup = kb)

    photo = open('image/team/bondarenko.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    bot.send_message(message.chat.id, text=text.olga_bondarenko, parse_mode='HTML')

    button_back_about_us (message)
#Контент - Достижения
@bot.message_handler(func=lambda message: message.text == "🥇Наші досягнення")
def achievements (message):
    bot.send_chat_action(message.chat.id, 'typing')
    photo_paths = text.img_invincibility
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    # media_group = []
    # for path in photo_paths:
    #     with open(path, "rb") as f:
    #         media_group.append(telebot.types.InputMediaPhoto(f.read()))
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.help_points_of_invincibility, parse_mode='HTML')

    photo_paths = text.img_donetsk
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.trip_to_donetsk_region, parse_mode='HTML')

    photo_paths = text.img_herson
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.herson, parse_mode='HTML')

    photo_paths = text.img_rana
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.rana, parse_mode='HTML')

    photo_paths = text.img_blessing_for_people
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.blessing_for_people, parse_mode='HTML')

    photo_paths = text.img_help_to_the_needy
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.help_to_the_needy, parse_mode='HTML')

    photo_paths = text.img_assistance_kherson_region
    media_group = [telebot.types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    bot.send_message(message.chat.id, text=text.assistance_kherson_region, parse_mode='HTML')
            
    button_back_about_us (message)
#Контент - Соц.сетях   
@bot.message_handler(func=lambda message: message.text == "💬Ми в соціальних мережах")
def social_networks (message):
    bot.send_chat_action(message.chat.id, 'typing')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Website', url='https://www.caringgeneration.org.ua')
    btn2= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/people/%D0%91%D0%BB%D0%B0%D0%B3%D0%BE%D0%B4%D1%96%D0%B9%D0%BD%D0%B0-%D0%9E%D1%80%D0%B3%D0%B0%D0%BD%D1%96%D0%B7%D0%B0%D1%86%D1%96%D1%8F-%D0%93%D1%80%D0%BE%D0%BC%D0%B0%D0%B4%D1%81%D1%8C%D0%BA%D0%B8%D0%B9-%D1%80%D1%83%D1%85-%D0%9F%D0%BE%D0%BA%D0%BE%D0%BB%D1%96%D0%BD%D0%BD%D1%8F-%D0%BD%D0%B5%D0%B1%D0%B0%D0%B9%D0%B4%D1%83%D0%B6%D0%B8%D1%85/100086522064845')
    btn3= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/caringgeneration_in_ua/')
    btn4= types.InlineKeyboardButton(text='Telegram channel', url='https://t.me/caringgeneration_in_ua')
    kb.add(btn1, btn2, btn3, btn4)
    bot.send_message(message.chat.id, "Соціальна мережа:", reply_markup = kb)
    button_back_about_us (message)
#------------конец------ Меню - про нас------

#Обработчик файлов
@bot.message_handler(content_types=['photo', 'document', 'voice', 'video'])
def handle_files(message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_ignor_no_text)
    else:
        bot.send_message(message.chat.id, text = text.ignor_no_text)

# Обработка сообщений
@bot.message_handler(content_types=['text'])
def word_processing(message):
    #В главное меню
    if message.text == "\u23EAВ головне меню" or message.text =="\u23EATo main menu":
        main_menu (message)
    #Назад с направления Про нас (все русские символы)
    elif message.text == "\U0001f519Назад":
        menu_about_us (message)
    #Назад с направления Оримати допомогу (первый символ анг. H)
    elif message.text == "\U0001f519Hазад":
        main_help_project (message)
    #Назад с направления Допомгти проекту (первый символ анг. а) (а - русская)
    elif message.text == "\U0001f519Нaзад" or message.text == "\U0001f519Bаck":
        main_menu_donats (message)
    #Назад с направления Монобанка,Банк.рах.,кріпта (первые 2-е на анг.)(Back - оригинал)
    elif message.text == "\U0001f519Haзад" or message.text == "\U0001f519Back":
        menu_money(message)
    #Назад с напрвления UAN,USD,EUR
    elif message.text == "\U0001f519Назaд" or message.text == "\U0001f519Baсk":
        bank_accounts(message)
    #Ответ на любой другой текст
    else:
        bot.send_chat_action(message.chat.id, 'typing')
        chat_id = message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(message.chat.id, text=text.eng_nezrozymiv, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, text=text.nezrozymiv, parse_mode='HTML')
        photo = open('image/nezrozymiv.jpg', 'rb')
        bot.send_photo(chat_id, photo)

bot.polling(non_stop=True)