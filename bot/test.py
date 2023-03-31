from dotenv import load_dotenv, find_dotenv
from telebot import TeleBot, logger, console_output_handler, types,apihelper
import logging
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
import text

load_dotenv(find_dotenv()) #подгрузить файл .env
bot = TeleBot(os.getenv('TEST_TOKEN')) #прочитать файл

#-- Запись логов
#Cоздаем объект класса Formatter
formatter = logging.Formatter('%(asctime)s (%(filename)s:%(lineno)d'+' %(threadName)s %(funcName)s) %(levelname)s - %(name)s: "%(message)s"',' %Y.%m.%d %H:%M:%S')
#Создаем обработчик для вывода сообщений лога в консоль
console_output_handler.setFormatter(formatter)
#Проверят создана ли папка logs
if not os.path.exists("logs"):
  os.mkdir("logs")
#Создаем обработчик для вывода сообщений лога в файл
fh = logging.FileHandler("logs/" + datetime.now().strftime(" %Y.%m.%d-%H.%M.%S") + ".log", encoding="utf-8")
fh.setFormatter(formatter)
logger.addHandler(fh)
#Задаем уровень логирования
logger.setLevel(logging.INFO)    
#-- Конец логов

#Проверят создан ли файл chatids.txt
if not os.path.exists("chatids.txt"):
    open("chatids.txt", "w").close()
#Загрузка массива с файла в SET
#Разсылка идет по SET, не по файлу (файл выступает в роли истории)
chatids_file = open("chatids.txt", "r")
chatids_users = set ()
for line in chatids_file:
    chatids_users.add(line.strip())
chatids_file.close()

#Создание словоря с iD и выбранным языком
user_languages = {}
#Команда /start - выбор языка
@bot.message_handler(commands=['start'])
def language_selection(message):
    #Если chat.id нет в сете (chatids_users) то добавляет
    if not str(message.chat.id) in chatids_users:
        chatids_users.add(message.chat.id)
        #Если chat.id нет в файле (chatids.txt) то добавляет
        if str(message.chat.id) not in open('chatids.txt').read():
            chatids_file = open("chatids.txt", "a")
            chatids_file.write(str(message.chat.id) + "\n")
            chatids_file.close()
    #Кнопки выбора языка
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text=text.language_selection)
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "🇺🇦Українська")
    btn2 = types.KeyboardButton(text ="🇬🇧English")
    kb.add(btn1, btn2,)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
    bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
#Отправка сообщение админом по команде /send
@bot.message_handler(commands=['send'])
def send_a_message(message):
    #Проверка на админа
    if message.chat.id in (759572442, 402411612):
        for user in chatids_users:
            bot.send_message(user, message.text[message.text.find(' '):])
    #если не админ выбивает стандартную (непонимайку)
    else:
        bot.send_chat_action(message.chat.id, 'typing')
        chat_id = message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(message.chat.id, text=text.eng_nezrozymiv, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, text=text.nezrozymiv, parse_mode='HTML')
        photo = open('image/nezrozymiv.jpg', 'rb')
        bot.send_photo(chat_id, photo)
# Функция для обработки выбора языка пользователем
@bot.message_handler(func=lambda message: message.text in ['🇬🇧English', '🇺🇦Українська'])
def language_preservation(message):
    #Cохраняем выбранный язык в словаре
    user_languages[message.chat.id] = message.text
    #Приветствие
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/start.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    main_menu (message)
# /share (Поделиться)
@bot.message_handler(commands=['share'])#Не переводиться на англ.
def share (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_zag_share, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_osn_share, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.zag_share, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.osn_share, parse_mode='HTML')
# Меню - Главное + команда
@bot.message_handler(commands=['menu']) #Не переводиться на англ.
def main_menu (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_privetstvie)
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f64fHelp the project")
        btn2 = types.KeyboardButton(text="\U0001faf6About us")
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        bot.send_message(message.chat.id, text = text.privetstvie)
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f198Отримати допомогу")
        btn2 = types.KeyboardButton(text ="\U0001f64fДопомогти проекту")
        btn3 = types.KeyboardButton(text ="\U0001f3ebОсвітні заходи")
        btn4 = types.KeyboardButton(text="\U0001faf6Про нас")
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)

#--Отправка сообщений   
class SendMessage:
    def __init__(self, message, button_location, button_who_are_you = None):
        self.message = message
        self.button_location = button_location
        self.button_who_are_you = button_who_are_you 

    def check_file(self):
        if not os.path.exists('Message.xlsx'):
            wb = Workbook()
            ws = wb.active
            ws.append(['yest_datetime', 'button_location', 'button_who_are_you', 'message.text', 'first_name', 'last_name', 'username', 'chat.id'])
            wb.save('Message.xlsx')
            
    def add_data(self):
        wb = load_workbook('Message.xlsx')
        sheet = wb.active
        last_row = sheet.max_row + 1
        yest_datetime = datetime.now()
        sheet.cell(row=last_row, column=1, value=yest_datetime)
        sheet.cell(row=last_row, column=2, value=self.button_location)
        sheet.cell(row=last_row, column=3, value=self.button_who_are_you)
        sheet.cell(row=last_row, column=4, value=self.message.text)
        sheet.cell(row=last_row, column=5, value=self.message.from_user.first_name)
        sheet.cell(row=last_row, column=6, value=self.message.from_user.last_name)
        sheet.cell(row=last_row, column=7, value=self.message.from_user.username)
        sheet.cell(row=last_row, column=8, value=self.message.chat.id)
        try:
            wb.save('Message.xlsx')
        except PermissionError:
            logger.exception("Не вдалось зберегти файл:")
            bot.send_chat_action(self.message.chat.id, 'typing')
            chat_id = self.message.chat.id
            if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
                bot.send_message(self.message.chat.id, text=text.eng_failed_to_send, parse_mode='HTML')
                bot.send_message(self.message.chat.id, text=text.eng_button_driver)
            else:
                bot.send_message(self.message.chat.id, text=text.failed_to_send, parse_mode='HTML')
                bot.send_message(self.message.chat.id, text=text.button_driver)
            return
        bot.send_chat_action(self.message.chat.id, 'typing')
        chat_id = self.message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(self.message.chat.id, text=text.eng_thank_contacting, parse_mode='HTML')
            bot.send_message(self.message.chat.id, text=text.eng_button_driver)
        else:
            bot.send_message(self.message.chat.id, text=text.thank_contacting, parse_mode='HTML')
            bot.send_message(self.message.chat.id, text=text.button_driver)
        if self.message.from_user.username:
            user = f"@{self.message.from_user.username}"
        else:
            user = f"{self.message.chat.id}"
        try:
            bot.send_message('-1001801043894', f"Графа: {self.button_location} \nВід користувача: {user}\nТекст повідомелння: \n{self.message.text}")
        except apihelper.ApiException:
            logger.exception("Дуже велике повідомлення, не вдалось відправити його в ТГ канал")
            return
    def process_request(self):
        self.check_file()
        self.add_data()
#Игнор кнопок
def ignor_button(message, button_location, button_who_are_you = None):
    if message.text == '/start':
        language_selection(message)
    elif message.text == '/menu':
        main_menu(message)
    elif message.text == '/share':
        share(message)
    elif message.text == '\U0001f519Hазад':
        main_help_project (message)
    elif message.text == '\U0001f519Haзaд':
        button_back_main_help_project (message)
    elif message.text == '\u23EAВ головне меню':
        main_menu(message)
    elif message.text == "\U0001f519Bаck":
        main_menu_donats (message)
    elif message.text == "\u23EATo main menu":
        main_menu(message)
    elif message.content_type != 'text' or len(message.text.split()) < 2:
        bot.send_chat_action(message.chat.id, 'typing')
        chat_id = message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(message.chat.id, text=text.eng_get_help_not_understand, parse_mode='HTML')
            bot.send_message(message.chat.id, text=text.eng_button_driver, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, text=text.get_help_not_understand, parse_mode='HTML')
            bot.send_message(message.chat.id, text=text.button_driver, parse_mode='HTML')
    else:
        help_zsy = SendMessage(message, button_location, button_who_are_you)
        help_zsy.process_request()
#--конец отправки сообщений
#------------ Меню Отримати допомогу-----
#Меню Отримати допомогу
@bot.message_handler(func=lambda message: message.text == "\U0001f198Отримати допомогу")
def main_help_project (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "👨‍⚖️Юридична консультація")
    btn2 = types.KeyboardButton(text = "📦Гуманітарна допомога")
    btn3 = types.KeyboardButton(text = "🙏Реалізувати Вашу мрію")
    btn4 = types.KeyboardButton(text = "\U0001fa96Допомога для ЗСУ")
    btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
    kb.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Юридическая консультация 
@bot.message_handler(func=lambda message: message.text == "👨‍⚖️Юридична консультація")
def legal_consultation(message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.legal_consultation, parse_mode='HTML')
    button_back_main_help_project (message)
#Меню Кто ТЫ
def button_back_main_help_project (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=1)
    btn1 = types.KeyboardButton(text = "\U0001fa96Я військовослужбовець")
    btn2 = types.KeyboardButton(text = "🧳Я внутрішньо переміщена особа")
    btn3 = types.KeyboardButton(text = "😢Я людина, яка постраждала від війни")
    btn4 = types.KeyboardButton(text = "🤲Я волонтер")
    btn5 = types.KeyboardButton(text = "\U0001f519Hазад")
    btn6 = types.KeyboardButton(text = "\u23EAВ головне меню")
    kb.add(btn1, btn2, btn3, btn4, btn5, btn6)
    bot.send_message(message.chat.id, "💁🏻‍♂️Тут ви можете отримати допомогу, але вам потрібно розповісти про себе", reply_markup=kb)
    bot.send_message(message.chat.id, text=text.button_driver)
#Я військовослужбовець
@bot.message_handler(func=lambda message: message.text == "\U0001fa96Я військовослужбовець")
def soldier (message):
    bot.send_chat_action(message.chat.id, 'typing')
    var_legal_consultation = "👨‍⚖️Юридична консультація"
    var_soldier = "\U0001fa96Я військовослужбовець"
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_legal_consultation, button_who_are_you = var_soldier)
    button_back_main_help_project_who_are_you (message)
#Я внутрішньо переміщена особа
@bot.message_handler(func=lambda message: message.text == "🧳Я внутрішньо переміщена особа")
def soldier (message):
    bot.send_chat_action(message.chat.id, 'typing')
    var_legal_consultation = "👨‍⚖️Юридична консультація"
    var_soldier = "🧳Я внутрішньо переміщена особа"
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_legal_consultation, button_who_are_you = var_soldier)
    button_back_main_help_project_who_are_you (message)
#Я людина, яка постраждала від війни
@bot.message_handler(func=lambda message: message.text == "😢Я людина, яка постраждала від війни")
def soldier (message):
    bot.send_chat_action(message.chat.id, 'typing')
    var_legal_consultation = "👨‍⚖️Юридична консультація"
    var_soldier = "😢Я людина, яка постраждала від війни"
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_legal_consultation, button_who_are_you = var_soldier)
    button_back_main_help_project_who_are_you (message)
#Я волонтер
@bot.message_handler(func=lambda message: message.text == "🤲Я волонтер")
def soldier (message):
    bot.send_chat_action(message.chat.id, 'typing')
    var_legal_consultation = "👨‍⚖️Юридична консультація"
    var_soldier = "🤲Я волонтер"
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_legal_consultation, button_who_are_you = var_soldier)
    button_back_main_help_project_who_are_you (message)
#Возвррат с выбора КТО ТЫ, в Юридическую консультацию
def button_back_main_help_project_who_are_you (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "\U0001f519Haзaд")
    btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")
    kb.add(btn1, btn2)
    bot.send_message(message.chat.id, text=text.button_driver, reply_markup=kb)
#Гуманітарна допомога
@bot.message_handler(func=lambda message: message.text == "📦Гуманітарна допомога")
def humanitarian_aid (message):
    bot.send_chat_action(message.chat.id, 'typing')
    var_humanitarian_aid = "📦Гуманітарна допомога"
    sent = bot.send_message(message.chat.id, text = text.your_situation, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_humanitarian_aid)
    button_back_main_help_project_locations (message)
#Допомога для ЗСУ
@bot.message_handler(func=lambda message: message.text == "\U0001fa96Допомога для ЗСУ")
def help_for_zsy(message):
    var_help_for_zsy = '\U0001fa96Допомога для ЗСУ'
    bot.send_chat_action(message.chat.id, 'typing')
    sent = bot.send_message(message.chat.id, text = text.help_zsy, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_help_for_zsy)
    button_back_main_help_project_locations (message)
#Реалізувати Вашу мрію
@bot.message_handler(func=lambda message: message.text == "🙏Реалізувати Вашу мрію")
def realize_your_dream(message):
    var_realize_your_dream = '🙏Реалізувати Вашу мрію'
    bot.send_chat_action(message.chat.id, 'typing')
    sent = bot.send_message(message.chat.id, text = text.realize_your_dream, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_realize_your_dream)
    button_back_main_help_project_locations (message)
#Возврат в Отримати допомогу
def button_back_main_help_project_locations (message):
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    btn1 = types.KeyboardButton(text = "\U0001f519Hазад")
    btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")
    kb.add(btn1, btn2)
    bot.send_message(message.chat.id, text=text.button_driver, reply_markup=kb)
#------------ конец----Меню Отримати допомогу-----

#------------ Меню Допомогти проекту-----
@bot.message_handler(func=lambda message: message.text == "\U0001f64fДопомогти проекту" or message.text =="\U0001f64fHelp the project")
def main_menu_donats (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "💰Financial assistance")
        btn2 = types.KeyboardButton(text = "🧦Other assistance")
        btn3 = types.KeyboardButton(text = "❓What will your help be aimed to")
        btn4 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "💰Фінансова допомога")
        btn2 = types.KeyboardButton(text = "🧦Інша допомога")
        btn3 = types.KeyboardButton(text = "❓На що буде спрямована ваша допомога")
        btn4 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата в Допомогти проекту
def menu_vozvrata_donats (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Bаck")
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Нaзад")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата в Монобанка,Банк.рах.,кріпта
def menu_vozvrata_mono_krypto (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Back")  
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Haзад")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню 💰Фінансова допомога-----
@bot.message_handler(func=lambda message: message.text == "💰Фінансова допомога" or message.text =="💰Financial assistance")
def menu_money (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001fad9Monobanka")
        btn2 = types.KeyboardButton(text = "💳Bank account")
        btn3 = types.KeyboardButton(text = "\U0001f519Bаck")
        btn4 = types.KeyboardButton(text = "\U0001fa99Cryptocurrency")
        btn5 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001fad9Monobanka")
        btn2 = types.KeyboardButton(text = "💳Банківські рахунки")
        btn3 = types.KeyboardButton(text = "\U0001f519Нaзад")
        btn4 = types.KeyboardButton(text = "\U0001fa99Криптовалюта")  
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню --Моно банка
@bot.message_handler(func=lambda message: message.text == "\U0001fad9Monobanka")
def menu_mono (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/monobanka.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='\U0001fad9Monobanka', url='https://send.monobank.ua/jar/AMrMtk2Vz')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text.eng_follow, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text.follow, reply_markup = kb)
    menu_vozvrata_mono_krypto (message)
#Меню 💳Банківські рахунки
@bot.message_handler(func=lambda message: message.text == "💳Банківські рахунки" or message.text =="💳Bank account")
def bank_accounts (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f1fa\U0001f1f8USD")
        btn2 = types.KeyboardButton(text = "\U0001f1ea\U0001f1faEUR")  
        btn3 = types.KeyboardButton(text = "\U0001f519Back")
        btn4 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f1fa\U0001f1e6UAN")
        btn2 = types.KeyboardButton(text = "\U0001f1fa\U0001f1f8USD")
        btn3 = types.KeyboardButton(text = "\U0001f519Haзад")
        btn4 = types.KeyboardButton(text = "\U0001f1ea\U0001f1faEUR")  
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Меню возврата с 💳Банківські рахунки
def menu_vozvrata_bank_accounts (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Baсk")  
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Назaд")  
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")  
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
# Кнопка UAN
@bot.message_handler(func=lambda message: message.text == "\U0001f1fa\U0001f1e6UAN")
def button_uan (message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.uan, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка USD
@bot.message_handler(func=lambda message: message.text == "\U0001f1fa\U0001f1f8USD")
def button_usd (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_usd, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.usd, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка EUR
@bot.message_handler(func=lambda message: message.text == "\U0001f1ea\U0001f1faEUR")
def button_eur (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_eur, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.eur, parse_mode='HTML')
    menu_vozvrata_bank_accounts (message)
#Кнопка - Криптовалюта
@bot.message_handler(func=lambda message: message.text == "\U0001fa99Криптовалюта" or message.text == "\U0001fa99Cryptocurrency")
def crypto (message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.USDT_crypto_check, parse_mode='HTML')
    bot.send_message(message.chat.id, text = text.BTC_crypto_check, parse_mode='HTML')
    menu_vozvrata_mono_krypto (message)
#Кнопка ❓На що буде спрямована ваша допомога
@bot.message_handler(func=lambda message: message.text == "❓На що буде спрямована ваша допомога" or message.text == "❓What will your help be aimed to")
def your_help_is_straightened (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_what_money_for, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text = text.what_money_for, parse_mode='HTML')
    menu_vozvrata_donats (message)
#🧦Інша допомога
@bot.message_handler(func=lambda message: message.text == "🧦Інша допомога" or message.text == "🧦Other assistance")
def other_help (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        var_legal_consultation = "🧦Other assistance"
        sent = bot.send_message(message.chat.id, text = text.eng_other_help_t, parse_mode='HTML')
    else:
        var_legal_consultation = "🧦Інша допомога"
        sent = bot.send_message(message.chat.id, text = text.other_help_t, parse_mode='HTML')
    bot.register_next_step_handler(sent, ignor_button, button_location=var_legal_consultation)
    menu_vozvrata_donats (message)
#------------ конец----Меню Допомогти проекту-----

#------------Меню Освітні заходи
@bot.message_handler(func=lambda message: message.text == "\U0001f3ebОсвітні заходи")
def educational_activities(message):
    bot.send_chat_action(message.chat.id, 'typing')
    bot.send_message(message.chat.id, text = text.educational_activities, parse_mode='HTML')
#------------ конец----Освітні заходи-----

#------------ Меню - Про нас
@bot.message_handler(func=lambda message: message.text == "\U0001faf6Про нас" or message.text == "\U0001faf6About us")
def menu_about_us (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "🧑‍💼Our founders")
        btn2 = types.KeyboardButton(text = "👪Our team")
        btn3 = types.KeyboardButton(text = "🥇Our achievements")
        btn4 = types.KeyboardButton(text = "💬We are on social networks")
        btn5 = types.KeyboardButton(text = "\u23EATo main menu")
        kb.add(btn1, btn2, btn3, btn4,btn5)
        bot.send_message(message.chat.id, text=text.eng_button_driver, reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "🧑‍💼Наші засновники")
        btn2 = types.KeyboardButton(text = "👪Наша команда")
        btn3 = types.KeyboardButton(text = "🥇Наші досягнення")
        btn4 = types.KeyboardButton(text = "💬Ми в соціальних мережах")
        btn5 = types.KeyboardButton(text = "\u23EAВ головне меню")
        kb.add(btn1, btn2, btn3, btn4,btn5)
        bot.send_message(message.chat.id, text=text.button_driver, reply_markup=kb)
#Кнопки возврата в меню: Про нас и Главное меню
def button_back_about_us (message):
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Bасk")
        btn2 = types.KeyboardButton(text = "\u23EATo main menu")
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.eng_button_driver,reply_markup=kb)
    else:
        kb = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
        btn1 = types.KeyboardButton(text = "\U0001f519Назад")
        btn2 = types.KeyboardButton(text = "\u23EAВ головне меню")
        kb.add(btn1, btn2)
        bot.send_message(message.chat.id, text=text.button_driver,reply_markup=kb)
#Контент - Основатели
@bot.message_handler(func=lambda message: message.text == "🧑‍💼Наші засновники" or message.text == "🧑‍💼Our founders")
def osnovatel (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/team/gatesh.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/vadim.gatezh')
    btn2= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/v_gatezh')
    btn3= types.InlineKeyboardButton(text='Telegram channel', url='https://t.me/v_gatezh_novyny')
    kb.add(btn1, btn2, btn3)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_vadim_gatezh, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.vadim_gatezh, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)
    
    photo = open('image/team/visotskiy.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/profile.php?id=100015205090408')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_pavlo_vysotsky, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.pavlo_vysotsky, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)
    button_back_about_us (message)
#Контент - Команда
@bot.message_handler(func=lambda message: message.text == "👪Наша команда" or message.text == "👪Our team")
def team (message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    photo = open('image/team/mironuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/anya.myroniuk')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_anna_mironyuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.anna_mironyuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)

    photo = open('image/team/merezhuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/merranst')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_anastasia_merezhuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.anastasia_merezhuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)

    photo = open('image/team/semenchuk.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/nastia_semenchuk')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_anastasia_semenchuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.anastasia_semenchuk, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)

    photo = open('image/team/birkova.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_julia_birkova, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.julia_birkova, parse_mode='HTML')

    photo = open('image/team/bagirov.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_vagif_bagirov, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.vagif_bagirov, parse_mode='HTML')

    photo = open('image/team/shaporda.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/shaporda.design')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_anastasia_shaporda, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.anastasia_shaporda, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)

    photo = open('image/team/torska.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/kristorska')
    kb.add(btn1)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_khrystyna_torska, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)        
    else:
        bot.send_message(message.chat.id, text=text.khrystyna_torska, parse_mode='HTML')
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)

    photo = open('image/team/bondarenko.jpg', 'rb')
    bot.send_photo(chat_id, photo)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_olga_bondarenko, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.olga_bondarenko, parse_mode='HTML')

    button_back_about_us (message)
#Контент - Достижения
@bot.message_handler(func=lambda message: message.text == "🥇Наші досягнення" or message.text == "🥇Our achievements")
def achievements (message):
    bot.send_chat_action(message.chat.id, 'typing')
    photo_paths = text.img_invincibility
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]

    bot.send_media_group(message.chat.id, media=media_group)
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_help_points_of_invincibility, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.help_points_of_invincibility, parse_mode='HTML')

    photo_paths = text.img_donetsk
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_trip_to_donetsk_region, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.trip_to_donetsk_region, parse_mode='HTML')

    photo_paths = text.img_herson
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_herson, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.herson, parse_mode='HTML')

    photo_paths = text.img_rana
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_rana, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.rana, parse_mode='HTML')

    photo_paths = text.img_blessing_for_people
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_blessing_for_people, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.blessing_for_people, parse_mode='HTML')

    photo_paths = text.img_help_to_the_needy
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
      bot.send_message(message.chat.id, text=text.eng_help_to_the_needy, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.help_to_the_needy, parse_mode='HTML')
        
    photo_paths = text.img_assistance_kherson_region
    media_group = [types.InputMediaPhoto(open(path, "rb").read()) for path in photo_paths]
    bot.send_media_group(message.chat.id, media=media_group)
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_assistance_kherson_region, parse_mode='HTML')
    else:
        bot.send_message(message.chat.id, text=text.assistance_kherson_region, parse_mode='HTML')

    button_back_about_us (message)
#Контент - Соц.сетях   
@bot.message_handler(func=lambda message: message.text == "💬Ми в соціальних мережах" or message.text == "💬We are on social networks")
def social_networks (message):
    bot.send_chat_action(message.chat.id, 'typing')
    kb = types.InlineKeyboardMarkup(row_width=1)
    btn1= types.InlineKeyboardButton(text='Website', url='https://www.caringgeneration.org.ua')
    btn2= types.InlineKeyboardButton(text='Facebook', url='https://www.facebook.com/people/%D0%91%D0%BB%D0%B0%D0%B3%D0%BE%D0%B4%D1%96%D0%B9%D0%BD%D0%B0-%D0%9E%D1%80%D0%B3%D0%B0%D0%BD%D1%96%D0%B7%D0%B0%D1%86%D1%96%D1%8F-%D0%93%D1%80%D0%BE%D0%BC%D0%B0%D0%B4%D1%81%D1%8C%D0%BA%D0%B8%D0%B9-%D1%80%D1%83%D1%85-%D0%9F%D0%BE%D0%BA%D0%BE%D0%BB%D1%96%D0%BD%D0%BD%D1%8F-%D0%BD%D0%B5%D0%B1%D0%B0%D0%B9%D0%B4%D1%83%D0%B6%D0%B8%D1%85/100086522064845')
    btn3= types.InlineKeyboardButton(text='Instagram', url='https://www.instagram.com/caringgeneration_in_ua/')
    btn4= types.InlineKeyboardButton(text='Telegram channel', url='https://t.me/caringgeneration_in_ua')
    kb.add(btn1, btn2, btn3, btn4)
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text=text.eng_social_network, reply_markup = kb)
    else:
        bot.send_message(message.chat.id, text=text.social_network, reply_markup = kb)
    button_back_about_us (message)
#------------конец------ Меню - про нас------

#Обработчик файлов
@bot.message_handler(content_types=['photo', 'document', 'voice', 'video', 'sticker', 'animation'])
def handle_files(message):
    bot.send_chat_action(message.chat.id, 'typing')
    chat_id = message.chat.id
    if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
        bot.send_message(message.chat.id, text = text.eng_ignor_no_text)
    else:
        bot.send_message(message.chat.id, text = text.ignor_no_text)

# Обработка сообщений
@bot.message_handler(content_types=['text'])
def word_processing(message):
    
    #В главное меню
    if message.text == "\u23EAВ головне меню" or message.text =="\u23EATo main menu":
        main_menu (message)
    #Назад с направления Про нас (все русские символы)
    elif message.text == "\U0001f519Назад" or message.text == "\U0001f519Bасk":
        menu_about_us (message)
    #Назад с направления Оримати допомогу (первый символ анг. H)
    elif message.text == "\U0001f519Hазад":
        main_help_project (message)
    #Назад с направления Юридическая консультация - КТО ТЫ
    elif message.text == "\U0001f519Haзaд":
        button_back_main_help_project (message)
    #Назад с направления Допомгти проекту (первый символ анг. а) (а - русская)
    elif message.text == "\U0001f519Нaзад" or message.text == "\U0001f519Bаck":
        main_menu_donats (message)
    #Назад с направления Монобанка,Банк.рах.,кріпта (первые 2-е на анг.)(Back - оригинал)
    elif message.text == "\U0001f519Haзад" or message.text == "\U0001f519Back":
        menu_money(message)
    #Назад с напрвления UAN,USD,EUR
    elif message.text == "\U0001f519Назaд" or message.text == "\U0001f519Baсk":
        bank_accounts(message)
    #Ответ на любой другой текст
    else:
        bot.send_chat_action(message.chat.id, 'typing')
        chat_id = message.chat.id
        if chat_id in user_languages and user_languages[chat_id] == '🇬🇧English':
            bot.send_message(message.chat.id, text=text.eng_nezrozymiv, parse_mode='HTML')
        else:
            bot.send_message(message.chat.id, text=text.nezrozymiv, parse_mode='HTML')
        photo = open('image/nezrozymiv.jpg', 'rb')
        bot.send_photo(chat_id, photo)

logger.info("---Запуск---")
try:
    bot.infinity_polling()
except Exception as e:
    logging.exception(e)